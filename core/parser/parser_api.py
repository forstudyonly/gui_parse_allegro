import json
import os
import glob
import pickle
import time
import requests
from openpyxl import Workbook, load_workbook
import re
from openpyxl.styles import Alignment
import core.check_api.tests
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service
from selenium import webdriver

other_words = ['V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO']

CLEANR = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')


def get_digits(s):
    try:
        return re.findall("(\d+\,\d+)", s)[0]
    except:
        return "-"


def refresh_token(dpg):
    dpg.set_value("status_api_text", "Обновляем токен...")
    link_val, code_verifier = core.check_api.tests.get_link()
    print(link_val)
    # webbrowser.open(link_val)
    print(code_verifier)
    chrome_service = Service(GeckoDriverManager().install())
    try:
        options = webdriver.FirefoxOptions()
        options.add_argument("--log-level=0")
        options.add_argument("--mute-audio")
        options.add_argument('no-sandbox')
        options.add_argument('--headless')
        with open("assets/firefox_location.txt", "r", encoding="utf-8") as file:
            bin_loc_firefox = file.read()
        options.binary_location = bin_loc_firefox
        try:
            with webdriver.Firefox(service=chrome_service, options=options) as driver:
                driver.set_page_load_timeout(10)
                driver.get("https://allegro.pl")
                with open("assets/cookies.txt", "r", encoding="utf-8") as file:
                    cookies = json.loads(file.read())
                for cook in cookies:
                    if cook.get("sameSite", False):
                        del cook["sameSite"]
                    driver.add_cookie(cook)
                driver.get(link_val)
                time.sleep(2)
                driver.close()
                driver.quit()
        except:
            print("poh")

    except Exception as ex:
        print(ex)

    time.sleep(1)
    with open("code.txt", "r", encoding="utf-8") as file:
        code_from_allegro = file.read()

    response = core.check_api.tests.get_access_token(code_from_allegro, code_verifier)
    print(response['access_token'])
    with open("access_token.txt", "w", encoding="utf-8") as file:
        file.write(response['access_token'])
    with open("access_token.txt", "r", encoding="utf-8") as file:
        access_token = file.read()
    res = core.check_api.tests.check_token(access_token)
    if res == 200:
        dpg.set_value("Output Text", "Токен рабочий!")

    elif res == 401:
        dpg.set_value("Output Text", "Токен устарел, нужно обновить!")
    else:
        dpg.set_value("Output Text", "Неизвестная ошибка")
    dpg.set_value("status_api_text", "Продолжаем")


def cleanhtml(raw_html):
    try:
        cleantext = re.sub(CLEANR, '', raw_html)
        return cleantext
    except BaseException as ex:
        return raw_html


def upload_imgs(images_urls, folder_name):
    names = []
    if not os.path.exists(f"data/{folder_name}/photos"):
        os.makedirs(f"data/{folder_name}/photos")
    for img_url in images_urls:
        try:
            if os.path.exists(f"data/{folder_name}/photos/{img_url.split('/')[-1]}.jpg"):
                names.append(f"data/{folder_name}/photos/{img_url.split('/')[-1]}.jpg")
                continue
            img = requests.get(img_url)
            with open(f"data/{folder_name}/photos/{img_url.split('/')[-1]}.jpg", "wb") as file:
                file.write(img.content)
            names.append(f"data/{folder_name}/photos/{img_url.split('/')[-1]}.jpg")
        except BaseException as ex:
            with open("error.txt", "a", encoding="utf-8") as file:
                file.write(f"\n{ex}\n")
            print(ex)
            continue
    return names


def get_product(productId, access_token, dpg):
    url = f"https://api.allegro.pl/sale/products/{productId}"
    headers = {
        "authorization": f"Bearer {access_token}",
        "accept": "application/vnd.allegro.public.v1+json"
    }
    while True:
        try:
            response = requests.get(url=url, headers=headers)
        except BaseException as ex:
            refresh_token()
            response = requests.get(url=url, headers=headers)
        # print(response.status_code)
        # print(response.text)
        if response.status_code == 200:
            json_book = json.loads(response.text)
            parameters = {}
            parameters["title"] = json_book["name"]
            parameters['other_info'] = []
            for param in json_book["parameters"]:
                if param["name"] == "ISBN":
                    parameters['isbn'] = param["values"][0]
                elif param["name"] == "Wydawnictwo":
                    parameters['izdatelstvo'] = param["valuesLabels"][0]
                elif param["name"] == "Autor":
                    parameters['avtor'] = param["valuesLabels"][0]
                elif param["name"] == "Okładka":
                    parameters['oblojka'] = param["valuesLabels"][0]
                elif param["name"] == "Tytuł":
                    parameters['titul'] = param["valuesLabels"][0]
                elif param["name"] == "Rok wydania":
                    parameters['god_izdaniya'] = param["valuesLabels"][0]
                elif param["name"] == "Liczba stron":
                    parameters['kolvo_stron'] = param["valuesLabels"][0]
                elif param["name"] == "Numer wydania":
                    parameters['num_wydania'] = param["valuesLabels"][0]
                elif param["name"] == "Szerokość produktu":
                    parameters['shirota'] = param["valuesLabels"][0]
                elif param["name"] == "Waga produktu z opakowaniem jednostkowym":
                    parameters['ves'] = param["valuesLabels"][0]
                elif param["name"] == "Język":
                    parameters['yazik'] = param["valuesLabels"][0]
                elif param['name'] == "Forma":
                    parameters['Forma'] = param["valuesLabels"][0]
                elif param['name'] == "Wysokość produktu":
                    parameters['visota'] = param["valuesLabels"][0]
                else:
                    parameters['other_info'] = parameters['other_info'] + [f"{param['name']}:{param['valuesLabels'][0]}"]
            imgs = []

            for img in json_book["images"]:
                imgs.append(img["url"])
            try:
                for item in json_book["description"]["sections"][0]["items"]:

                    if item["type"] == "TEXT":
                        parameters["desc"] = item["content"]
                        break
            except BaseException as ex:
                print(ex)
                with open("error.txt", "a", encoding="utf-8") as file_error:
                    file_error.write(f"\n{ex}\n")
                parameters["desc"] = ""
            parameters["images"] = imgs
            break
        else:
            refresh_token()
            continue
    return parameters


with open("assets/all_categories.data", "rb") as file:
    all_categories = pickle.load(file)


def get_category_value(category_path, access_token):
    category_value = []
    try:
        for categoryId in category_path:
            check_exist_cat = all_categories.get(f"{categoryId}", False)
            if check_exist_cat:
                category_value.append(check_exist_cat)
                continue

            url = f"https://api.allegro.pl/sale/categories/{categoryId}"
            headers = {
                "authorization": f"Bearer {access_token}",
                "accept": "application/vnd.allegro.public.v1+json"
            }
            response = requests.get(url=url, headers=headers)
            #print(response.text)
            json_val = json.loads(response.text)
            category_value.append(json_val["name"])
            all_categories[categoryId] = json_val["name"]
        return category_value
    except BaseException as ex:
        with open("error.txt", "a", encoding="utf-8") as file_error:
            file_error.write(f"\n{ex}\n")
        return []


def start_api_parse(folder_name, dpg):
    try:
        refresh_token(dpg)
        dpg.set_value("status_api_text", "Начали работу...")
        with open("access_token.txt", "r", encoding="utf-8") as file_token:
            access_token = file_token.read()
        # получаем все папки для парса
        all_dirs = [i.replace(os.getcwd(), "").replace("/data", "") for i in
        glob.glob(os.getcwd() + f"/data/{folder_name}/*")]
        all_dirs = [i[i.find(folder_name) + len(folder_name):].replace("//", "").replace("\\", "") for i in all_dirs]
        start_time = time.time()
        all_dirs = [i if ".xlsx" not in i and "photos" != i else "" for i in all_dirs]
        print(all_dirs)
        for dir in all_dirs:
            if dir == "":
                continue
            # Получаем все файлы
            all_files_for_dir = glob.glob(os.getcwd() + f"/data/{folder_name}/{dir}/*")
            dpg.set_value("status_api_text", f"Парсим {dir} категорию")
            dpg.set_value("Progress Bar API", 0)
            count = 2
            if not os.path.exists(f'data/{folder_name}/{dir}.xlsx'):
                wb = Workbook()
                ws = wb.active
                start_file_value = 0
                ws[f'A1'] = "ID товара"
                ws[f'B1'] = "ISBN"
                ws[f'C1'] = "Полное название"
                ws[f'D1'] = "Price"
                ws[f'E1'] = "Продано шт."
                ws[f'F1'] = "Wydawnictwo"
                ws[f'G1'] = "Автор"
                ws[f'H1'] = "Szerokość produktu"
                ws[f'I1'] = "Wysokość produktu"
                ws[f'J1'] = "Waga produktu z opakowaniem jednostkowym"
                ws[f'K1'] = "Categories"
                ws[f'L1'] = "Titul"
                ws[f'M1'] = "Okładka"
                ws[f'N1'] = "Rok wydania"
                ws[f'O1'] = "Liczba stron"
                ws[f'P1'] = "Numer wydania"
                ws[f'Q1'] = "Język"
                ws[f'R1'] = "Forma"
                ws["S1"] = "Описание"
                ws["T1"] = "imgs"
                ws["U1"] = "Ссылка"
                ws["V1"] = "Остальные категории формата название:значение"

                wb.save(f'data/{folder_name}/{dir}.xlsx')
                wb.close()
            else:
                file_path = f'data/{folder_name}/{dir}.xlsx'
                workbook = load_workbook(file_path)
                sheet = workbook.active
                num_rows = int(sheet.max_row)
                # Выводим количество строк на экран
                print("Количество строк: ", num_rows)
                if float((num_rows - 1) / 60) > len(all_files_for_dir) - 1:
                    continue
                start_file_value = int((num_rows - 1) / 60)
                count = start_file_value * 60 + 2
                workbook.close()
                print(f"start_file_value - {start_file_value}")
                print(f"count - {count}")
                dpg.set_value("Progress Bar API", (start_file_value * 60) / (len(all_files_for_dir) * 60))


            #for page in all_files_for_dir:
            try:
                wb = load_workbook(f'data/{folder_name}/{dir}.xlsx')
                ws = wb.active
                for i in range(start_file_value, len(all_files_for_dir)):
                    page = os.getcwd() + f"/data/{folder_name}/{dir}/data{i}.json"
                    print(page)
                    # Для каждой стр начинаем парс
                    with open(page, "r", encoding="utf-8") as file_json:
                        page_info = file_json.read()
                    json_data = json.loads(page_info)
                    listings = json.loads(json_data['__listing_StoreState'])
                    all_params = []
                    for element in listings["items"]["elements"]:
                        if element["type"] == "label":
                            continue

                        # json_item = json.dumps(element)
                        # print(json_item)
                        try:
                            price = element["price"]["main"]["amount"]
                            url = element["url"]

                            prod_id = element["productId"]
                            categories = element["assortmentCategory"]["path"]
                            osob_kupilo = element.get("popularityLabel", "0")
                            if osob_kupilo != "0":
                                osob_kupilo = osob_kupilo.split(" ")[0]
                        except BaseException as ex:
                            with open("error.txt", "a", encoding="utf-8") as file_error:
                                file_error.write(f"\n{ex}\n")
                            continue
                        params = get_product(prod_id, access_token, dpg)
                        params["osob_kupilo"] = osob_kupilo
                        params["prod_id"] = prod_id
                        params["categories"] = categories
                        params["price"] = price
                        params["url"] = url
                        all_params.append(params)
                        try:
                            val = dpg.get_value("Progress Bar API")
                            dpg.set_value("Progress Bar API", val + 1 / (len(all_files_for_dir) * 60))
                        except BaseException as ex:
                            with open("error.txt", "a", encoding="utf-8") as file_error:
                                file_error.write(f"\n{ex}\n")
                            continue
                    try:
                        dpg.set_value("status_api_text", f"Сейчас закрывать программу не рекомендуется, идет запись информации и файл может повредится!")
                    except BaseException as ex:
                        with open("error.txt", "a", encoding="utf-8") as file_error:
                            file_error.write(f"\n{ex}\n")
                        print(ex)
                    for params in all_params:
                        try:
                            ws[f'A{count}'] = params.get("prod_id", "-")
                            ws[f'B{count}'] = params.get("isbn", "-")
                            ws[f'C{count}'] = params.get("title", "-")

                            ws[f'D{count}'] = params.get("price", "-").replace(".", ",")
                            ws[f'D{count}'].number_format = '# ###0,00 zł'
                            ws[f'D{count}'].alignment = Alignment(horizontal='center')

                            ws[f'E{count}'] = int(params.get("osob_kupilo", "0"))
                            ws[f'E{count}'].alignment = Alignment(horizontal='center')

                            ws[f'F{count}'] = params.get("izdatelstvo", "-")
                            ws[f'G{count}'] = params.get("avtor", "-")

                            ws[f'H{count}'] = get_digits(params.get("shirota", "-"))
                            ws[f'H{count}'].alignment = Alignment(horizontal='center')
                            ws[f'I{count}'] = get_digits(params.get("visota", "-"))
                            ws[f'I{count}'].alignment = Alignment(horizontal='center')
                            ws[f'J{count}'] = get_digits(params.get("ves", "-"))
                            ws[f'J{count}'].alignment = Alignment(horizontal='center')

                            ws[f'K{count}'] = " - ".join(get_category_value(params["categories"], access_token))

                            ws[f'L{count}'] = params.get("titul", "-")
                            ws[f'M{count}'] = params.get("oblojka", "-")
                            ws[f'N{count}'] = params.get("god_izdaniya", "-")
                            ws[f'O{count}'] = params.get("kolvo_stron", "-")
                            ws[f'P{count}'] = params.get("num_wydania", "-")
                            ws[f'Q{count}'] = params.get("yazik", "-")
                            ws[f'R{count}'] = params.get("Forma", "-")
                            ws[f'S{count}'] = cleanhtml(params.get("desc", "-"))
                            imgs_paths = upload_imgs(params.get("images"), folder_name)
                            ws[f'T{count}'] = "\n".join(imgs_paths)
                            ws[f'U{count}'] = params.get("url", "-")

                            num_other = 0

                            if len(params["other_info"]) > 0:
                                for param in params["other_info"]:
                                    if num_other > len(other_words) - 1:
                                        break
                                    word_value = other_words[num_other]
                                    ws[f'{word_value}{count}'] = param
                                    num_other += 1
                            wb.save(f'data/{folder_name}/{dir}.xlsx')
                            # print(f"Спарсили {count - 1} книгу")
                            count += 1
                            # val = dpg.get_value("Progress Bar API")
                            # dpg.set_value("Progress Bar API", val + 1 / (len(all_files_for_dir) * 60))
                        except BaseException as ex:
                            with open("error.txt", "a", encoding="utf-8") as file:
                                file.write(f"\n{ex}\n")
                            continue
                    dpg.set_value("status_api_text", f"Парсим {dir} категорию")
            except BaseException as ex:
                with open("error.txt", "a", encoding="utf-8") as file:
                    file.write(f"\n{ex}\n")
                print(ex)
            finally:
                try:
                    wb.close()
                except BaseException as ex:
                    with open("error.txt", "a", encoding="utf-8") as file_error:
                        file_error.write(f"\n{ex}\nExcel уже закрыт")
                    print("Excel уже закрыт")


            dpg.set_value("Progress Bar API", 1)

        with open("assets/all_categories.data", "wb") as file:
            pickle.dump(all_categories, file)
        print(f"Заняло времени: {time.time() - start_time}")
        dpg.set_value("status_api_text", "Готово!")
    except BaseException as ex:
        with open("error.txt", "a", encoding="utf-8") as file_error:
            file_error.write(f"\n{ex}\n")
        dpg.set_value("status_api_text", "Какая-то ошибка, сокрее всего устарел API токен!")
        print(ex)
